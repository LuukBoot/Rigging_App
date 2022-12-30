from pathlib import Path 
from openpyxl import load_workbook



# Loading factors DNV 
Constans_DNV_folder_path = Path(__file__).parent
Constans_DNV_file_path = Constans_DNV_folder_path.parent / "Data_input/Constants_DNV.xlsx"

wb_factors_dnv = load_workbook(Constans_DNV_file_path)
Fact_names = wb_factors_dnv.sheetnames

Fact_text = {}
Fact_values = {}
for factor in Fact_names:
    # Getting worksheet 
    ws=wb_factors_dnv[factor]
    # Starting at row 2 see excel
    i=2
    text_values = []
    text="a"
    row_factors = {}

    while text!=None:
        text=ws["A"+str(i)].value
        if text!=None:
            # Daf is other type of data storage see Excel
            if factor == "DAF":
                text_values = ["Offshore","Inshore"]
                row_factors[text] = [ws["B"+str(i)].value,
                                    ws["C"+str(i)].value]
            else:
                row_factors[text] = ws["B"+str(i)].value
                
                text_values.append(text)
            i+=1
    Fact_values[factor] = row_factors
    Fact_text[factor] = text_values

# Import data from excel sheet 
Rigging_data_File = Path(__file__).parent
Rigging_data_File_Path = Rigging_data_File.parent / "Data_input/Data_lifting_tools.xlsx"
wb_Rigging_data = load_workbook(Rigging_data_File_Path, data_only=True)
ws_Rigging_data = wb_Rigging_data["Hijsmateriaal 1.4"]
# Number of rows in the rigging app file
number_of_row_file = 2000
id_list_total = []
df_rigging= {"Slings" : {},
             "Shackles": {},
             "Other":{},
             "drop_list_connect": ["Hook",
                                    "Pin",
                                    "Trunnion",
                                    "Shackle not in data base"]}
for i in  range(1,number_of_row_file):
    
    id = str(ws_Rigging_data["AJ"+str(i)].value)

    if id != None and id not in id_list_total:
        id_list_total.append(id)
        # Determine type equipment
        type_eqp = str(ws_Rigging_data["F"+str(i)].value).lower()

        if "grommet" in type_eqp or "sling" in type_eqp:
            info = {}
            info["type"] = ws_Rigging_data["F"+str(i)].value
            info["Length"] = ws_Rigging_data["G"+str(i)].value
            info["Dia"] = ws_Rigging_data["H"+str(i)].value
            info["MBL"] = ws_Rigging_data["J"+str(i)].value
            df_rigging["Slings"].update({id:info})

        elif "shackle" in type_eqp:
            txt = "Shackle: " + str(id)
            df_rigging["drop_list_connect"].append(txt)
            info = {}
            info["WLL"]=ws_Rigging_data["M"+str(i)].value
            info["D"]=ws_Rigging_data["O"+str(i)].value
            df_rigging["Shackles"].update({id:info})
        else: 
            info = {}
            info["WLL"]=ws_Rigging_data["M"+str(i)].value
            info["Weight"]=ws_Rigging_data["AC"+str(i)].value
            info["Type"]=type_eqp
            df_rigging["Other"].update({id:info})





        
# Default tables

Table_Lift_points=[{"point":"A","x":0,"y":0,"z":0},
                    {"point":"B","x":0,"y":0,"z":0},
                    {"point":"C","x":0,"y":0,"z":0},
                    {"point":"D","x":0,"y":0,"z":0} ]

Table_slings_grommets = [{"ID_number":"Type id number",
                          "type":"sling",
                          "SWL":1,
                          "D":1,
                          "n_parts":1}]
Table_connect =[{"side":"Upper","connections":'Hook',"dia":1,"SWL":1},
                {"side":"Lower","connections":'Hook',"dia":1,"SWL":1},
]
# Functions for visualtiaztion
def get_point_options(params,**kwargs):
    point_names = ["A","B","C","D"]
    N_lifts = params["N_lifts"]
    _point_names = point_names[0:N_lifts]
   
    
    return _point_names

def get_SSF_options(params,**kwargs):
    options = []
    name="SSF"
    i=1
    for row in params.tab_2.section_2["SSF_data"]:
        options.append(name+str(i))
        i+=1
    return options